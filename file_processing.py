import os
import aiofiles
import csv
import textract
from typing import List
from fastapi import UploadFile, HTTPException
from docx import Document
from openpyxl import load_workbook  # ✅ Import for XLSX processing
from utils import extract_text_from_pdf
from supabase import create_client, Client  # ✅ Import Supabase

# Supabase Configuration
SUPABASE_URL = "https://tkpimadtgruwryrwsofm.supabase.co"  # Replace with your Supabase project URL
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InRrcGltYWR0Z3J1d3J5cndzb2ZtIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NDA0MzQ3MTgsImV4cCI6MjA1NjAxMDcxOH0.W-KWnVcClcWWcbX_x1EdtmN2V_Jq-6wr1rMWxlbByVM"  # Replace with your Supabase service role key
SUPABASE_BUCKET = "uploads"  # Replace with your bucket name

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

async def upload_files(files: List[UploadFile]) -> dict:
    """
    Uploads files to Supabase Storage and extracts text.
    Supports PDF, TXT, DOC, DOCX, PAGES, RTF, MD, CSV, and XLSX.
    """
    combined_text = ""
    uploaded_files = []

    for file in files:
        try:
            file_extension = file.filename.split(".")[-1].lower()
            file_path = f"uploads/{file.filename}"  # Path inside Supabase Storage

            # ✅ Read file content as bytes
            file_content = await file.read()

            # ✅ Upload file to Supabase
            response = supabase.storage.from_(SUPABASE_BUCKET).upload(
                file_path,
                file_content,
                file_options={"contentType": file.content_type}
            )

            if "error" in response:
                raise HTTPException(status_code=500, detail=f"Failed to upload {file.filename} to Supabase: {response['error']}")

            # ✅ Generate public URL for the uploaded file
            file_url = f"{SUPABASE_URL}/storage/v1/object/public/{SUPABASE_BUCKET}/{file_path}"
            uploaded_files.append({"filename": file.filename, "url": file_url})

            # ✅ Extract text if applicable
            if file_extension == "pdf":
                pdf_text = extract_text_from_pdf(file_url)  # Modify function to support URLs if necessary
                combined_text += pdf_text + "\n"

            elif file_extension in ["txt", "md"]:
                combined_text += file_content.decode("utf-8") + "\n"

            elif file_extension == "docx":
                # Read DOCX from a temporary file instead of a URL
                with open("/tmp/temp.docx", "wb") as temp_file:
                    temp_file.write(file_content)
                doc = Document("/tmp/temp.docx")
                combined_text += "\n".join([p.text for p in doc.paragraphs]) + "\n"

            elif file_extension in ["doc", "pages", "rtf"]:
                with open("/tmp/tempfile", "wb") as temp_file:
                    temp_file.write(file_content)
                textract_text = textract.process("/tmp/tempfile").decode("utf-8")
                combined_text += textract_text + "\n"

            elif file_extension == "csv":
                csv_reader = csv.reader(file_content.decode("utf-8").splitlines())
                combined_text += "\n".join([", ".join(row) for row in csv_reader]) + "\n"

            elif file_extension == "xlsx":
                with open("/tmp/temp.xlsx", "wb") as temp_file:
                    temp_file.write(file_content)
                wb = load_workbook("/tmp/temp.xlsx", data_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        combined_text += ", ".join(str(cell) if cell else "" for cell in row) + "\n"

        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error processing {file.filename}: {str(e)}")

    return {
        "message": f"Files uploaded successfully",
        "files": uploaded_files,
        "extracted_text": combined_text,
    }